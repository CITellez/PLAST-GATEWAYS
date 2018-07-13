from openpyxl import *
import time
import sys  

#reload(sys)  
#sys.setdefaultencoding('utf8')

FILE_PATH = 'C:\Users\uidh3600\Desktop\E8A34000.xlsx'
file = open('C:\Users\uidh3600\Desktop\hello.dbc', 'r')
wb = load_workbook(FILE_PATH, read_only=True)

document = []

'''Guardamos solo lo que nos importa en una lista llamada document'''
for line in file:
    if line.startswith(' SG_'):
        document.append(line)
size = len(document)
print size


Chanel = {'C':1, 'B':2, 1:3, 2:4, 3:5}
           
for sheet in wb:
    print(sheet.title)    
    '''Creamos un arreglo que nos va a ayudar a guardar el nombre de las hojas de excel'''
    time.sleep(.1)

    for num in range(3,18):
        sender_chanel       = Chanel[sheet.cell(row = num, column = 1).value]
        receiver_chanel     = Chanel[sheet.cell(row = num, column = 2).value]
        sender_msg          = sheet.cell(row = num, column = 4).value
        sen_sig             = sheet.cell(row = num, column = 5).value
        sen_sig             = sen_sig.encode("utf-8")
        receiver_msg        = sheet.cell(row = num, column = 7).value
        receiver_sig        = sheet.cell(row = num, column = 8).value
        requirement         = sheet.cell(row = num, column = 12).value

        print 'from chanel' , sender_chanel, 'to chanel', receiver_chanel
        print sender_msg,'::',sen_sig,'--- to ---',receiver_msg,'::',receiver_sig
        
        for num in range(0,size):
            if document[num].find(sen_sig) != -1:
                print document[num]
                signal = document[num]
                sizee = len(signal)
                print sizee
                for numm in range(0,sizee):
                    if signal[numm] == "@":
                        print signal[numm+1]
                        print signal[numm-1]
                        desde = int(signal[numm+1])
                        hasta = int(signal[numm-1])

                        total = (2**hasta)
                        print total

                           
        print 'can_instance.SetSignalValue(sender_chanel,sender_msg,sen_sig,9)'
        time.sleep(.1)
        print 'can_instance.SetSignalValue(sender_chanel,sender_msg,sen_sig,9)'
        time.sleep(.1)
        print 'can_instance.SetSignalValue(sender_chanel,sender_msg,sen_sig,9)'
        time.sleep(.1)
        print 'can_instance.SetSignalValue(sender_chanel,sender_msg,sen_sig,9)'
        time.sleep(.1)
        print 'can_instance.SetSignalValue(sender_chanel,sender_msg,sen_sig,9)'
        time.sleep(.1)

        print '''response_receiver = can_instance.GetSignalValue(receiver_chanel, reciver_msg ,receiver_sig)'''
        time.sleep(.1)
        print '''response_receiver = can_instance.GetSignalValue(receiver_chanel, reciver_msg ,receiver_sig)'''
        time.sleep(.1)
        print '''response_sender = can_instance.GetSignalValue(receiver_chanel, reciver_msg ,receiver_sig)'''
        time.sleep(.1)
        print '''response_sender = can_instance.GetSignalValue(receiver_chanel, reciver_msg ,receiver_sig)\n'''
        time.sleep(.1)       
'''
        try:
            assert(response_receiver == response_sender)
        except
            print 'El GATEWAY',sender_msg,'::',sen_sig,'--- to ---',receiver_msg,'::',receiver_sig,'NO FUNCIONA'
            
'''
