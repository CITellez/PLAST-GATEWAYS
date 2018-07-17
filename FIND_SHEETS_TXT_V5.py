from openpyxl import *
import time
import sys  

#reload(sys)  
#sys.setdefaultencoding('utf8')

#////////////////////////////////////////////////////////////////////////////////////////////////
#////////////////////////
#////////////////////////////////////////////////////////////////////////////////////////////////
def save_for(total):
    string_total = str(total)

    results.write('    for send in range(0,')
    results.write(string_total)
    results.write('):\n')

def save_send(schanel,smessage,ssignal):
    for i in range(5):
        string_schanel  = str(schanel)
        string_smessage = str(smessage)
        string_ssignal  = str(ssignal)
        
        results.write('        can_instance.SetSignalValue(')
        results.write(string_schanel)
        results.write(',')
        results.write(string_smessage)
        results.write(',')
        results.write(string_ssignal)
        results.write(',send )\n')
        results.write('        time.sleep(.1)\n')
    results.write('\n')

def save_receive(rchanel,rmessage,rsignal):
    for i in range(5):
        string_schanel  = str(rchanel)
        string_rmessage = str(rmessage)
        string_rsignal  = str(rsignal)

        results.write('        response_receiver = can_instance.GetSignalValue(')
        results.write(string_schanel)
        results.write(',')
        results.write(string_rmessage)
        results.write(',')
        results.write(rsignal)
        results.write(')\n')
        results.write('        time.sleep(.1)\n')
    results.write('\n')

#//////////////////////////////////////////////////////////////////////////////////////////////////
#////////////////////// Open files, excel, databases and .py   ////////////////////////////////////
#//////////////////////////////////////////////////////////////////////////////////////////////////
try:
    FILE_PATH = 'D:\Python\E8A34000.xlsx'
    wb = load_workbook(FILE_PATH, read_only=True)
except:
    print 'El excel donde estan los requerimientos no esta el la direccion especificada'
    
try:
    file_CANC = open('D:\Python\PDT_E3A_R2_CCAN_Hand_Edited_Version2.dbc', 'r')
except:
    print 'La dbc de can C no se encuentra en la direccion especifica o el nombre no es correcto'
    
try:
    file_CANB = open('D:\Python\PDT20_E1A_R1_BHCAN.dbc', 'r')
except:
    print 'La dbc de can B no se encuentra en la direccion especifica o el nombre no es correcto'

try:
    results = open("results.txt","w+")
    results.write('hello')
    #results.close()
    #se esta guardando en C:\Users\uidh3600
except:
    print 'El documento de resultados no puedo ser creado'

#//////////////////////////////////////////////////////////////////////////////////////////////////
#//////////////////////// Initial code (it doesn't change)/////////////////////////////////////////
#//////////////////////////////////////////////////////////////////////////////////////////////////
x = 'soy una variable de prueba :V '
results.write('# Autor: Carlos Ichiro Tellez Urakami\n')
results.write('# Purpose: Gateways automatization\n')
results.write('# Date:\n')
results.write('#/////////////////////////////////////\n\n')
results.write('from PlatformFiles import *\n')
results.write('from _UserFunctions import *\n')
results.write('from openpyxl import *\n')
results.write('import time\n')

results.write('def Gateways(testcase):\n\n')
results.write('    testcase.safety_requirement = ,\n')  #Necesitas ver como poner un caracter '' dentro
results.write('    testcase.traceability = ,\n')             
results.write('    can_instance = CANoe()\n')
results.write('    try:\n')
results.write('        can_instance.Start()\n')
results.write('    except:\n')
results.write('        print("La simulacion ya esta corriendo")\n')
#results.close()

#//////////////////////////////////////////////////////////////////////////////////////////////////
#////////////////////////      Variables to be used      //////////////////////////////////////////
#//////////////////////////////////////////////////////////////////////////////////////////////////
document_canc = []
document_canb = []
tatal = 0
desde = 0
hasta = 0
Chanel = {'C':1, 'B':2, 1:3, 2:4, 3:5}

#//////////////////////////////////////////////////////////////////////////////////////////////////
#/////////////////////Extracting the importan information from the DBCs ///////////////////////////
#//////////////////////////////////////////////////////////////////////////////////////////////////

for line in file_CANC:
    if line.startswith(' SG_'):
        document_canc.append(line)
size = len(document_canc) 
print size

for line in file_CANB:
    if line.startswith(' SG_'):
        document_canb.append(line)
size = len(document_canb) 
print size

#//////////////////////////////////////////////////////////////////////////////////////////////////
#/////////////////////     Start working with the excel information     ///////////////////////////
#//////////////////////////////////////////////////////////////////////////////////////////////////          
for sheet in wb:
    print(sheet.title)    
    time.sleep(.1)

    for num in range(3,18):
        sender_chanel       = Chanel[sheet.cell(row = num, column = 1).value]
        receiver_chanel     = Chanel[sheet.cell(row = num, column = 2).value]
        sender_msg          = sheet.cell(row = num, column = 4).value
        sen_sig             = sheet.cell(row = num, column = 5).value
        sen_sig             = sen_sig.encode("utf-8")
        receiver_msg        = sheet.cell(row = num, column = 7).value
        receiver_sig        = sheet.cell(row = num, column = 8).value
        requirement         = sheet.cell(row = num, column = 12).value

        print 'from chanel' , sender_chanel, 'to chanel', receiver_chanel
        print sender_msg,'::',sen_sig,'--- to ---',receiver_msg,'::',receiver_sig
        time.sleep(1)
        #identify the databese to be used 
        for signal_line in document_canc:
            if signal_line.find(sen_sig) != -1:
                print signal_line
                signal = signal_line #what is returning 
                sizee = len(signal) 
                print sizee
                indexDot = signal.find("@")
                
                desde = signal[indexDot+1]
                hasta = int(signal[indexDot-1])
        total = (2**hasta)
        print total

        save_for(total)
        save_send(sender_chanel,sender_msg,sen_sig)
        save_receive(receiver_chanel,receiver_msg,receiver_sig)
    results.close()
    '''
        try:
            assert(response_receiver == response_sender)
        except
            print 'El GATEWAY',sender_msg,'::',sen_sig,'--- to ---',receiver_msg,'::',receiver_sig,'NO FUNCIONA'
            
    '''