'''
    Author: L. Nitu
    Purpose: --
    Date: 26-04-2016        
    Arguments: -   
    Outputs: -    
    Dependencies: -

    History:
    ------------------------------------------------------------------------------
    Revision | Date:      | Modification:                             | Author
    .......................................................................
    v0.1     | 26-04-2016 | Creation                                  | L. Nitu 
    ------------------------------------------------------------------------------
'''

from PlatformFiles import *
from _UserFunctions import *
from openpyxl import *
# from ParserDBC_Interface import dbc_parser, can_message 
# from PlatformFiles.CANOE_Interface import CANoe
#    can_message
import time
from Tix import COLUMN
from _sqlite3 import Row

 

def TEST_ExS1_1(testcase):
    
    print "Esta es mi primera prueba gracias por su comprencion"
    testcase.safety_requirement = 'Yes'
    testcase.traceability = 'REQ_ExS1_1'
    testcase.description = \
    """ 
    When the ECU is awake the cycle_func shall be executed every 100ms.
    """
#     testcase.skip('Skip reason: equipments not present to run this test.')
    
#     cMgr = ic.ConnectionMgr()  # create the winIDEA connection
#     cMgr.connectMRU('')  # initialize winIDEA
#     dbg = ic.CDebugFacade(cMgr)  # get winIDEA instance
    
    testcase.addPrecondition('stop target')
#     dbg.stop()
    
    testcase.addPrecondition('delete all breakpoints')
#     dbg.deleteAll()
    
    testcase.addPrecondition('reset target')
#     dbg.reset()
    
    testcase.addStep(1, 'Set breakpoint to cycle_func function')
#     dbg.setBP('cycle_func')
    
    testcase.addStep(1, 'run target')
#     dbg.run()
    
    testcase.addStep(1, 'wait until stopped')
#     dbg.waitUntilStopped()
    
    testcase.addStep(1, 'Evaluate first_time = value of STM System Timer Module (address: 0xFFF3C004)')
#     firstTime = dbg.evaluate(IConnectDebug.fRealTime, '*(uint32*)(0xFFF3C004)').getLong()
    
    testcase.addStep(2, 'run target')
#     dbg.run()
    
    testcase.addStep(2, 'wait until stopped')
#     dbg.waitUntilStopped()
    
    testcase.addStep(2, 'Evaluate second_time = value of STM System Timer Module (address: 0xFFF3C004)')
#     secondTime = dbg.evaluate(IConnectDebug.fRealTime, '*(uint32*)(0xFFF3C004)').getLong()
    
    testcase.addExpectedResult(2, '(first_time - second_time) <= 100')
#     assert(((secondTime - firstTime) / 1000) <= 100)
    
#     dbg.deleteAll()
#     dbg.run()
    

def TEST_ExS1_2(testcase):
    
    testcase.safety_requirement = 'No'
    testcase.traceability = 'REQ_ExS1_2'
    testcase.description = \
    """ 
    Write here the requirement or what this test case cover.
    You can write on multiple lines.
    """
    
    assert(1 == 2)
        
def TEST_ExS1_3(testcase):  
    
    testcase.safety_requirement = 'No'
    testcase.traceability = 'REQ_ExS1_3'
    testcase.description = \
    """ 
    This case will be used to create the first test case made by me 
    
    mucho ojo el can que esta funcionando es el can C
    """
    can_instance = CANoe()
    
    """       The test works althoug CANoe is not running
       //////////////////////////////////////////////////////"""
    try:    
        can_instance.Start()
    except:
        print("La simulacion ya esta corriendo")   
    
#     can_instance.SetSignalValue(2,'ESP_A4', 'VehAccel_Y', 200)
    print("El automovil ha sido encendido")
    
    """To read something we need a certain time between write and read
       ////////////////////////////////////////////////////////////////"""    
    time.sleep(.1)
    print can_instance.GetSignalValue(1, 'CBC_CFG1' ,'RKE_ONE_PRESS_ENBL')
    
    FILE_PATH = 'C:\Users\uidh3600\Desktop\GATEWAYS_MYJT20.xlsx'

    SHEET = 'CAN to CAN'

    workbook = load_workbook(FILE_PATH, read_only=True)

    sheet = workbook[SHEET]
    
    #for num in range(3,176):
    message_sender  = sheet.cell(row = 126, column = 4).value
    signal_sender = sheet.cell(row = 126, column = 5).value
    
    print can_instance.SignalFullName(2,message_sender,signal_sender)
    time.sleep(.1)
    can_instance.SetSignalValue(2,message_sender,signal_sender,9)
    time.sleep(.5)
    can_instance.SetSignalValue(2,message_sender,signal_sender,9)
    time.sleep(.5)
    can_instance.SetSignalValue(2,message_sender,signal_sender,9)
    time.sleep(.5)
    can_instance.SetSignalValue(2,message_sender,signal_sender,9)
    time.sleep(.5)
    
    print(message_sender) 
    print(signal_sender) 
    
   
               
    #can_instance.Stop()
    
    
    
    assert(1 == 1)
    
def TEST_ExS1_4(testcase):
    
    testcase.safety_requirement = 'No'
    testcase.traceability = 'REQ_ExS1_4'
    testcase.description = \
    """ 
    Write here the specification or what this test case cover.
    You can write on multiple lines.
    """
    
    can_instance = CANoe()
    can_instance.Start()
    time.sleep(5)
#     can_instance.Stop()
    can_instance.SetSignalValue(1, 'RFHUB_A1', 'IgnPos', 1)
    time.sleep(5)
    
    
    
    
    assert(1 == 1)  # error left intentionally
