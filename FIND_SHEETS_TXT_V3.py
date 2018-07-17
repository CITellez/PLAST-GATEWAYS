from openpyxl import *
import time
import sys  

#reload(sys)  
#sys.setdefaultencoding('utf8')
try:
    FILE_PATH = 'D:\Python\E8A34000.xlsx'
    wb = load_workbook(FILE_PATH, read_only=True)
except:
    print 'El excel donde estan los requerimientos no esta el la direccion especificada'
    
try:
    file_CANC = open('D:\Python\PDT_E3A_R2_CCAN_Hand_Edited_Version2.dbc', 'r')
except:
    print 'La dbc de can C no se encuentra en la direccion especifica o el nombre no es correcto'
    
try:
    file_CANB = open('D:\Python\PDT20_E1A_R1_BHCAN.dbc', 'r')
except:
    print 'La dbc de can B no se encuentra en la direccion especifica o el nombre no es correcto'


# Document is a list variable used to storage the lines of the file that are usefull 
document_canc = []
document_canb = []
tatal = 0
desde = 0
hasta = 0
Chanel = {'C':1, 'B':2, 1:3, 2:4, 3:5}



# Every line that starts with 'SG_' will be append in the document variable 
for line in file_CANC:
    if line.startswith(' SG_'):
        document_canc.append(line)
size = len(document_canc) 
print size

#''' Every line that starts with 'SG_' will be append in the document variable '''
for line in file_CANB:
    if line.startswith(' SG_'):
        document_canb.append(line)
size = len(document_canb) 
print size


#''' Every worksheet in the workbook will be open '''           
for sheet in wb:
    print(sheet.title)    
    time.sleep(.1)

    for num in range(3,18):
        sender_chanel       = Chanel[sheet.cell(row = num, column = 1).value]
        receiver_chanel     = Chanel[sheet.cell(row = num, column = 2).value]
        sender_msg          = sheet.cell(row = num, column = 4).value
        sen_sig             = sheet.cell(row = num, column = 5).value
        sen_sig             = sen_sig.encode("utf-8")
        receiver_msg        = sheet.cell(row = num, column = 7).value
        receiver_sig        = sheet.cell(row = num, column = 8).value
        requirement         = sheet.cell(row = num, column = 12).value

        print 'from chanel' , sender_chanel, 'to chanel', receiver_chanel
        print sender_msg,'::',sen_sig,'--- to ---',receiver_msg,'::',receiver_sig
        time.sleep(1)
        for signal_line in document_canc:
            if signal_line.find(sen_sig) != -1:
                print signal_line
                signal = signal_line #what is returning 
                sizee = len(signal) 
                print sizee
                indexDot = signal.find("@")
                
                desde = signal[indexDot+1]
                hasta = int(signal[indexDot-1])
        total = (2**hasta)
        print total

        print 'for send in range(0,',total,'):'                    
        print 'can_instance.SetSignalValue(', sender_chanel , ',' , sender_msg, ',' ,sen_sig, ',' ,'send' ,')'
        time.sleep(.1)
        print 'can_instance.SetSignalValue(', sender_chanel , ',' , sender_msg, ',' ,sen_sig, ',' ,'send' ,')'
        time.sleep(.1)
        print 'can_instance.SetSignalValue(', sender_chanel , ',' , sender_msg, ',' ,sen_sig, ',' ,'send' ,')'
        time.sleep(.1)
        print 'can_instance.SetSignalValue(', sender_chanel , ',' , sender_msg, ',' ,sen_sig, ',' ,'send' ,')'
        time.sleep(.1)
        print 'can_instance.SetSignalValue(', sender_chanel , ',' , sender_msg, ',' ,sen_sig, ',' ,'send' ,')'
        time.sleep(.1)

        print 'response_receiver = can_instance.GetSignalValue(',receiver_chanel,',',receiver_msg,',',receiver_sig,')'
        time.sleep(.1)
        print 'response_receiver = can_instance.GetSignalValue(',receiver_chanel,',',receiver_msg,',',receiver_sig,')'
        time.sleep(.1)
        print 'response_receiver = can_instance.GetSignalValue(',receiver_chanel,',',receiver_msg,',',receiver_sig,')'
        time.sleep(.1)
        print 'response_receiver = can_instance.GetSignalValue(',receiver_chanel,',',receiver_msg,',',receiver_sig,')'
        time.sleep(.1)       
    '''
        try:
            assert(response_receiver == response_sender)
        except
            print 'El GATEWAY',sender_msg,'::',sen_sig,'--- to ---',receiver_msg,'::',receiver_sig,'NO FUNCIONA'
            
    '''
