from openpyxl import *
import time

FILE_PATH = 'C:\Users\uidh3600\Desktop\GATEWAYS_MYJT20.xlsx'
wb = load_workbook(FILE_PATH, read_only=True)

Chanel = {'C':1, 'B':2, 1:3, 2:4, 3:5}
           
for sheet in wb:
    print(sheet.title)    
    '''Creamos un arreglo que nos va a ayudar a guardar el nombre de las hojas de excel'''
    time.sleep(1)

    for num in range(3,18):
        sender_chanel       = Chanel[sheet.cell(row = num, column = 1).value]
        receiver_chanel     = Chanel[sheet.cell(row = num, column = 2).value]
        sender_msg          = sheet.cell(row = num, column = 4).value
        sender_sig          = sheet.cell(row = num, column = 5).value
        receiver_msg        = sheet.cell(row = num, column = 7).value
        receiver_sig        = sheet.cell(row = num, column = 8).value
        requirement         = sheet.cell(row = num, column = 12).value

        print 'from chanel' , sender_chanel, 'to chanel', receiver_chanel
        print sender_msg,'::',sender_sig,'--- to ---',receiver_msg,'::',receiver_sig

        

        print 'can_instance.SetSignalValue(sender_chanel,sender_msg,sender_sig,9)'
        time.sleep(.2)
        print 'can_instance.SetSignalValue(sender_chanel,sender_msg,sender_sig,9)'
        time.sleep(.2)
        print 'can_instance.SetSignalValue(sender_chanel,sender_msg,sender_sig,9)'
        time.sleep(.2)
        print 'can_instance.SetSignalValue(sender_chanel,sender_msg,sender_sig,9)'
        time.sleep(.2)
        print 'can_instance.SetSignalValue(sender_chanel,sender_msg,sender_sig,9)'
        time.sleep(1)

        print '''response_receiver = can_instance.GetSignalValue(receiver_chanel, reciver_msg ,receiver_sig)'''
        time.sleep(.1)
        print '''response_receiver = can_instance.GetSignalValue(receiver_chanel, reciver_msg ,receiver_sig)'''
        time.sleep(.1)
        print '''response_sender = can_instance.GetSignalValue(receiver_chanel, reciver_msg ,receiver_sig)'''
        time.sleep(.1)
        print '''response_sender = can_instance.GetSignalValue(receiver_chanel, reciver_msg ,receiver_sig)'''
        time.sleep(.1)
'''
        try:
            assert(response_receiver == response_sender)
        except
            print 'El GATEWAY',sender_msg,'::',sender_sig,'--- to ---',receiver_msg,'::',receiver_sig,'NO FUNCIONA'
            
'''
        

        

        

        
        
            
        






    

